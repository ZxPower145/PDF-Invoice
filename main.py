import pandas as pd
import openpyxl
from fpdf import FPDF
import glob
from pathlib import Path

# Get everything with an .xlsx extension
filepaths = glob.glob("invoices/*.xlsx")

# Read evey Excel file in the filepaths
for filepath in filepaths:
    pdf = FPDF(orientation="L", unit="mm", format="A4")

    pdf.add_page()
# Get the filename from each file, and also get the invoice number and the date from the file name
    file_names = Path(filepath).stem
    invoice_number, date = file_names.split("-")

    pdf.set_font("Times", style="B", size=16)
    pdf.cell(h=8, w=50, align="L", txt=f"Invoice nr. {invoice_number}", ln=1)
    pdf.cell(h=8, w=50, align="L", txt=f"Date: {date}", ln=1)

    pdf.ln(15)

    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    pdf.set_font("Times", style="B", size=10)

# Add a header

    columns = df.columns
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.cell(w=30, h=12, txt=columns[0], border=1)
    pdf.cell(w=70, h=12, txt=columns[1], border=1)
    pdf.cell(w=40, h=12, txt=columns[2], border=1)
    pdf.cell(w=30, h=12, txt=columns[3], border=1)
    pdf.cell(w=30, h=12, txt=columns[4], border=1, ln=1)

# Add rows

    for index, row in df.iterrows():
        pdf.set_font("Times", size=10)
        pdf.cell(w=30, h=12, txt=str(row['product_id']), border=1)
        pdf.cell(w=70, h=12, txt=str(row['product_name']), border=1)
        pdf.cell(w=40, h=12, txt=str(row['amount_purchased']), border=1)
        pdf.cell(w=30, h=12, txt=str(row['price_per_unit']), border=1)
        pdf.cell(w=30, h=12, txt=str(row['total_price']), border=1, ln=1)

# Add a total sum

    total_sum = str(df["total_price"].sum())

    pdf.set_font("Times", size=10)
    pdf.cell(w=30, h=12, txt="", border=1)
    pdf.cell(w=70, h=12, txt="", border=1)
    pdf.cell(w=40, h=12, txt="", border=1)
    pdf.cell(w=30, h=12, txt="", border=1)
    pdf.cell(w=30, h=12, txt=total_sum, border=1, ln=1)

    pdf.cell(w=0, h=12, txt="", ln=1)

# Add total price sentence:
    pdf.set_font("Times", size=16, style="B")
    pdf.cell(w=0, h=12, txt=f"Total price is: {total_sum} EUR", ln=1)

# Add company name and logo

    pdf.cell(w=25, h=12, txt=f"ZxPower")
    pdf.image("img.png", w=10)

    pdf.output(f"PDFs/Invoice {invoice_number}.pdf")
